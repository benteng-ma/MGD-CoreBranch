#!/usr/bin/env python
# -*- coding: utf-8 -*-

"""
Cross-model directional concordance between Hsd3b6-KO and Awat2-KO.

This script:

1. Reads the author-provided Hsd3b6 supplementary RNA-seq table.
2. Defines Hsd3b6 FDR-upregulated genes as:
       log2FoldChange > 0 and adj_pvalue < 0.05
3. Builds whole-MG pseudobulk CPM profiles for WT1, WT2, WT3 and KO
   from the frozen Awat2 high-confidence MG raw-count object.
4. Defines reference-robust Awat2 directional elevation as:
       KO CPM > WT1 CPM, WT2 CPM and WT3 CPM
5. Tests whether independently defined Hsd3b6 FDR-UP genes are enriched
   for this Awat2 KO > all-WT pattern using a one-sided Fisher exact test.
6. Reports the odds ratio and 95% confidence interval.

Important:
Awat2 contains only one pooled KO library, so this is a directional
cross-model concordance analysis, not replicate-level differential
expression testing.
"""

from pathlib import Path

import numpy as np
import openpyxl
import pandas as pd
import scanpy as sc

from scipy.stats import fisher_exact
from scipy.stats.contingency import odds_ratio


# ------------------------------------------------------------
# Project paths
# ------------------------------------------------------------

PROJECT_ROOT = Path(__file__).resolve().parents[1]

AWAT2_FILE = (
    PROJECT_ROOT
    / "source_data"
    / "GSE261036"
    / "GSE261036_MG_highconfidence_rawcounts.h5ad"
)

HSD_FILE = (
    PROJECT_ROOT
    / "source_data"
    / "GSE166784"
    / "Supplementary_Data_1.xls"
)

OUT_DIR = PROJECT_ROOT / "results" / "cross_model"
OUT_DIR.mkdir(parents=True, exist_ok=True)


# ------------------------------------------------------------
# 1. Read Hsd3b6 author supplementary table
# ------------------------------------------------------------

print("Reading Hsd3b6 author supplement...")

# The file has an .xls extension, but it is readable by openpyxl
# when opened through a binary file handle.
#
# IMPORTANT:
# With read_only=True, sequential iter_rows() must be used.
# Repeated ws.cell() access is extremely slow for this workbook.

with open(HSD_FILE, "rb") as fh:

    wb = openpyxl.load_workbook(
        fh,
        data_only=True,
        read_only=True,
    )

    if "RNAseq" not in wb.sheetnames:
        raise ValueError(
            "RNAseq sheet was not found in Supplementary_Data_1.xls"
        )

    ws = wb["RNAseq"]

    row_iter = ws.iter_rows(values_only=True)

    try:
        header_row = next(row_iter)
    except StopIteration:
        raise ValueError(
            "RNAseq sheet is empty."
        )

    headers = {
        str(name): i
        for i, name in enumerate(header_row)
        if name is not None
    }

    required_columns = {
        "gene",
        "log2FoldChange",
        "pvalue",
        "adj_pvalue",
    }

    missing_columns = required_columns.difference(headers)

    if missing_columns:
        raise ValueError(
            f"Missing required Hsd3b6 columns: {sorted(missing_columns)}"
        )

    rows = []

    for values in row_iter:

        gene = values[headers["gene"]]
        log2fc = values[headers["log2FoldChange"]]
        pvalue = values[headers["pvalue"]]
        adj_pvalue = values[headers["adj_pvalue"]]

        if gene is None:
            continue

        if log2fc is None or adj_pvalue is None:
            continue

        rows.append(
            {
                "gene": str(gene),
                "log2FoldChange": float(log2fc),
                "pvalue": (
                    float(pvalue)
                    if pvalue is not None
                    else np.nan
                ),
                "adj_pvalue": float(adj_pvalue),
            }
        )

    wb.close()


hsd = pd.DataFrame(rows)

if hsd.empty:
    raise ValueError(
        "No Hsd3b6 RNA-seq rows were successfully read."
    )


# ------------------------------------------------------------
# 2. Define Hsd3b6 FDR-UP genes
# ------------------------------------------------------------

hsd_up = (
    hsd[
        (hsd["log2FoldChange"] > 0)
        & (hsd["adj_pvalue"] < 0.05)
    ]
    .copy()
    .sort_values(
        ["adj_pvalue", "gene"],
        ascending=[True, True],
    )
    .reset_index(drop=True)
)

hsd_up.to_csv(
    OUT_DIR
    / "Hsd3b6_FDR_UP_from_author_supplement.csv",
    index=False,
)

print("Hsd3b6 FDR-UP genes:", len(hsd_up))


# ------------------------------------------------------------
# 3. Read frozen Awat2 whole-MG raw-count object
# ------------------------------------------------------------

print("Reading frozen Awat2 whole-MG raw counts...")

adata = sc.read_h5ad(AWAT2_FILE)

if "sample" not in adata.obs.columns:
    raise ValueError(
        "Awat2 AnnData is missing obs['sample']."
    )

if "gene_symbol" not in adata.var.columns:
    raise ValueError(
        "Awat2 AnnData is missing var['gene_symbol']."
    )

samples = [
    "WT1",
    "WT2",
    "WT3",
    "KO",
]

observed_samples = set(
    adata.obs["sample"]
    .astype(str)
)

missing_samples = [
    sample
    for sample in samples
    if sample not in observed_samples
]

if missing_samples:
    raise ValueError(
        f"Missing Awat2 libraries: {missing_samples}"
    )


# ------------------------------------------------------------
# 4. Whole-MG sample-level pseudobulk CPM
# ------------------------------------------------------------

gene_symbols = (
    adata.var["gene_symbol"]
    .astype(str)
    .to_numpy()
)

pseudobulk = {
    "gene": gene_symbols
}

for sample in samples:

    mask = (
        adata.obs["sample"]
        .astype(str)
        .eq(sample)
        .to_numpy()
    )

    n_cells = int(mask.sum())

    if n_cells == 0:
        raise ValueError(
            f"No cells found for sample {sample}"
        )

    counts = np.asarray(
        adata.X[mask, :].sum(axis=0)
    ).ravel().astype(float)

    library_size = counts.sum()

    if library_size <= 0:
        raise ValueError(
            f"Zero library size for sample {sample}"
        )

    pseudobulk[sample] = (
        counts
        / library_size
        * 1e6
    )


pb = pd.DataFrame(pseudobulk)

# Remove invalid symbols.
pb = pb[
    ~pb["gene"].isin(
        [
            "",
            "nan",
            "None",
        ]
    )
].copy()

# Collapse duplicated gene symbols by summing CPM.
pb = (
    pb
    .groupby(
        "gene",
        as_index=False,
    )[samples]
    .sum()
)

# Formal expressed-background definition:
# gene must reach >= 1 CPM in at least one of the four libraries.
pb = pb[
    pb[samples].max(axis=1) >= 1
].copy()

# Reference-robust Awat2 directional criterion.
pb["KO_above_allWT"] = (
    pb["KO"]
    >
    pb[
        [
            "WT1",
            "WT2",
            "WT3",
        ]
    ].max(axis=1)
)

pb = (
    pb
    .sort_values("gene")
    .reset_index(drop=True)
)

pb.to_csv(
    OUT_DIR
    / "Awat2_wholeMG_allGene_pseudobulk_CPM.csv",
    index=False,
)


# ------------------------------------------------------------
# 5. Intersect Hsd3b6 FDR-UP genes with Awat2 background
# ------------------------------------------------------------

hsd_gene_set = set(
    hsd_up["gene"]
)

detected = (
    pb[
        pb["gene"].isin(
            hsd_gene_set
        )
    ]
    .copy()
    .sort_values("gene")
    .reset_index(drop=True)
)

detected.to_csv(
    OUT_DIR
    / "Hsd3b6_FDR_UP_detected_in_Awat2.csv",
    index=False,
)


# ------------------------------------------------------------
# 6. Fisher exact enrichment
# ------------------------------------------------------------

panel_yes = int(
    detected["KO_above_allWT"].sum()
)

panel_no = int(
    len(detected)
    - panel_yes
)

background_yes = int(
    pb["KO_above_allWT"].sum()
)

background_no = int(
    len(pb)
    - background_yes
)

# Construct mutually exclusive 2x2 table:
#
#                    KO > all WT    not KO > all WT
# Hsd3b6 FDR-UP          a                b
# all other genes        c                d

remainder_yes = (
    background_yes
    - panel_yes
)

remainder_no = (
    background_no
    - panel_no
)

table = [
    [
        panel_yes,
        panel_no,
    ],
    [
        remainder_yes,
        remainder_no,
    ],
]

fisher_or, fisher_p = fisher_exact(
    table,
    alternative="greater",
)

conditional_or_result = odds_ratio(
    table,
    kind="conditional",
)

ci95 = (
    conditional_or_result
    .confidence_interval(
        confidence_level=0.95
    )
)


# ------------------------------------------------------------
# 7. Save formal summary
# ------------------------------------------------------------

summary = pd.DataFrame(
    [
        {
            "background_n": len(pb),
            "background_KO_above_allWT": background_yes,
            "background_fraction": (
                background_yes
                / len(pb)
            ),
            "Hsd3b6_FDR_UP_total": len(hsd_up),
            "Hsd3b6_FDR_UP_detected": len(detected),
            "Hsd3b6_FDR_UP_KO_above_allWT": panel_yes,
            "Hsd3b6_fraction": (
                panel_yes
                / len(detected)
                if len(detected) > 0
                else np.nan
            ),
            "Fisher_OR": fisher_or,
            "one_sided_P": fisher_p,
            "conditional_OR": (
                conditional_or_result.statistic
            ),
            "CI95_low": ci95.low,
            "CI95_high": ci95.high,
        }
    ]
)

summary.to_csv(
    OUT_DIR
    / "Hsd3b6_Awat2_directional_concordance_summary.csv",
    index=False,
)


# ------------------------------------------------------------
# 8. Console report
# ------------------------------------------------------------

supporting_genes = (
    detected.loc[
        detected["KO_above_allWT"],
        "gene",
    ]
    .tolist()
)

non_supporting_genes = (
    detected.loc[
        ~detected["KO_above_allWT"],
        "gene",
    ]
    .tolist()
)

print()

print(
    "EXPRESSED BACKGROUND:",
    len(pb),
)

print(
    "BACKGROUND KO>allWT:",
    background_yes,
    "/",
    len(pb),
    "=",
    round(
        background_yes / len(pb),
        4,
    ),
)

print()

print(
    "Hsd3b6 FDR-UP total:",
    len(hsd_up),
)

print(
    "Hsd3b6 FDR-UP detected:",
    len(detected),
)

print(
    "Hsd3b6-UP KO>allWT:",
    panel_yes,
    "/",
    len(detected),
    "=",
    round(
        panel_yes / len(detected),
        4,
    ),
)

print(
    "SUPPORTING GENES:",
    ",".join(supporting_genes),
)

print(
    "NON-SUPPORTING GENES:",
    ",".join(non_supporting_genes),
)

print()

print(
    "2x2 TABLE:",
    table,
)

print(
    "FISHER OR =",
    fisher_or,
)

print(
    "ONE-SIDED P =",
    fisher_p,
)

print(
    "CONDITIONAL OR =",
    conditional_or_result.statistic,
)

print(
    "95% CI =",
    ci95.low,
    ci95.high,
)

print()

print("DONE")

print(
    "Results saved to:",
    OUT_DIR,
)